import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, fills, Alignment, Side, Border, PatternFill, fonts
#from conexionBD import *

class crear_estado_de_cuenta():
    def __init__(self) -> None:
        pass
    def crear(self, nombre, fecha_inicio, cantidad_pagos, modal, monto, cuota):

        #Otras variables
        retornar = cantidad_pagos * cuota
        ganancia = retornar - monto
        moneda = '0,000.00$'
        letras = ['B', 'C', 'D', 'E', 'F', 'G']
        palabra_celda = ""
        #Fin

        #Colores HEXADECIMALES
        CREMA = 'FFE599'
        OSCURO = '000000'
        BLANCO = 'FFFFFF'
        AMARILLO = 'FBBC04'
        AZUL = '6FA8DC'
        VERDE = '38761D'
        ROJO = '980000'
        ROSA = 'D5A6BD'
        #Fin

        #Creacion de la hoja
        arch = openpyxl.Workbook()
        hoja = arch.active
        hoja.title = nombre
        #Fin

        #Unir celdas
        hoja.merge_cells('B3:C3')
        hoja.merge_cells('D7:D8')
        hoja.merge_cells('E7:E8')
        hoja.merge_cells('F5:G8')
        #Fin

        #Celdas
        celda_nombre = hoja['B3']

        celda_fecha = hoja['B4']
        celda_fecha_dato = hoja['C4']
        celda_gana = hoja['D4']
        celda_gana_dato = hoja['E4']

        celda_monto = hoja['B5']
        celda_monto_dato = hoja['C5']
        celda_cuota = hoja['D5']
        celda_cuota_dato = hoja['E5']

        celda_cantidad_pagos = hoja['B6']
        celda_cantidad_pagos_dato = hoja['C6']
        celda_modalidad = hoja['D6']
        celda_modalidad_dato = hoja['E6']

        celda_cantidad_realizados = hoja['B7']
        celda_cantidad_realizados_dato = hoja['C7']

        celda_retornar = hoja['D7']
        celda_retornar_dato = hoja['E7']

        celda_cantidad_pendiente = hoja['B8']
        celda_cantidad_pendiente_dato = hoja['C8']

        celda_gerome = hoja['F5']

        celda_numero = hoja['B9']
        celda_cuota_escrita = hoja['C9']
        celda_fecha_escrita = hoja['D9']
        celda_pagado_escrita = hoja['E9']
        celda_balance_escrita = hoja['F9']
        celda_comentario_escrita = hoja['G9']
        #Fin

        #Estilo de las celdas
        font_color_negro_18 = Font(color = OSCURO, name = 'Arial', size = 18, bold = True)
        font_color_blanco_18 = Font(color = BLANCO, name = 'Arial', size = 18, bold = True)
        font_color_negro_18_nobold = Font(color = OSCURO, name = 'Arial', size = 18, bold = False)
        font_color_blanco_18_nobold = Font(color = BLANCO, name = 'Arial', size = 18, bold = False)
        font_color_blanco_24 = Font(color = BLANCO, name = 'Arial', size = 24, bold = True)
        font_color_negro_24 = Font(color = OSCURO, name = 'Arial', size = 24, bold = True)
        font_color_rojo_24 = Font(color = ROJO, name = 'Arial', size = 24, bold = True)
        font_color_negro_12 = Font(color = OSCURO, name = 'Arial', size = 12, bold = True)
        font_color_negro_12_nobold = Font(color = OSCURO, name = 'Arial', size = 12, bold = False)
        
        fill_color_crema = PatternFill(fill_type = fills.FILL_SOLID, start_color = CREMA)
        fill_color_amarillo = PatternFill(fill_type = fills.FILL_SOLID, start_color = AMARILLO)
        fill_color_verde = PatternFill(fill_type = fills.FILL_SOLID, start_color = VERDE)
        fill_color_azul = PatternFill(fill_type = fills.FILL_SOLID, start_color = AZUL)
        fill_color_rosa = PatternFill(fill_type = fills.FILL_SOLID, start_color = ROSA)

        border = Border(left = Side (border_style = 'thin', color = OSCURO), right = Side (border_style = 'thin', color = OSCURO), top = Side (border_style = 'thin', color = OSCURO), bottom = Side (border_style = 'thin', color = OSCURO))

        aliniacion_centro_abajo = Alignment(horizontal="center", vertical="bottom", wrap_text = True)
        aliniacion_derecha = Alignment(horizontal="right", vertical="center")
        aliniacion_central = Alignment(horizontal="center", vertical="center")

        hoja.column_dimensions['B'].width = 38
        hoja.column_dimensions['C'].width = 35
        hoja.column_dimensions['D'].width = 20
        hoja.column_dimensions['E'].width = 25
        hoja.column_dimensions['F'].width = 20
        hoja.column_dimensions['G'].width = 26
        #Fin

        #Dar estilo a las celdas
        celda_nombre.font = font_color_negro_18
        celda_nombre.fill = fill_color_crema
        celda_nombre.border = border

        celda_fecha.font = font_color_blanco_18
        celda_fecha.fill = fill_color_azul
        celda_fecha.border = border
        celda_fecha.alignment = aliniacion_derecha

        celda_fecha_dato.font = font_color_blanco_18_nobold
        celda_fecha_dato.fill = fill_color_azul
        celda_fecha_dato.border = border
        celda_fecha_dato.alignment = aliniacion_central

        celda_gana.font = font_color_negro_18
        celda_gana.fill = fill_color_crema
        celda_gana.border = border

        celda_gana_dato.font = font_color_negro_18
        celda_gana_dato.fill = fill_color_crema
        celda_gana_dato.border = border
        celda_gana_dato.alignment = aliniacion_derecha
        celda_gana_dato.number_format = moneda

        celda_monto.font = font_color_negro_18
        celda_monto.border = border
        celda_monto.alignment = aliniacion_derecha

        celda_monto_dato.font = font_color_blanco_24
        celda_monto_dato.fill = fill_color_verde
        celda_monto_dato.border = border
        celda_monto_dato.alignment = aliniacion_central
        celda_monto_dato.number_format = moneda

        celda_cuota.font = font_color_negro_18_nobold
        celda_cuota.border = border
        celda_cuota.alignment = aliniacion_central

        celda_cuota_dato.font = font_color_blanco_24
        celda_cuota_dato.fill = fill_color_verde
        celda_cuota_dato.border = border
        celda_cuota_dato.alignment = aliniacion_central
        celda_cuota_dato.number_format = moneda

        celda_cantidad_pagos.font = font_color_negro_18
        celda_cantidad_pagos.border = border
        celda_cantidad_pagos.alignment = aliniacion_derecha

        celda_cantidad_pagos_dato.font = font_color_negro_24
        celda_cantidad_pagos_dato.border = border
        celda_cantidad_pagos_dato.alignment = aliniacion_central

        celda_modalidad.font = font_color_negro_18_nobold
        celda_modalidad.fill = fill_color_crema
        celda_modalidad.border = border
        celda_modalidad.alignment = aliniacion_central

        celda_modalidad_dato.font = font_color_negro_18_nobold
        celda_modalidad_dato.fill = fill_color_crema
        celda_modalidad_dato.border = border
        celda_modalidad_dato.alignment = aliniacion_central

        celda_cantidad_realizados.font = font_color_negro_18
        celda_cantidad_realizados.fill = fill_color_azul
        celda_cantidad_realizados.border = border
        celda_cantidad_realizados.alignment = aliniacion_derecha

        celda_cantidad_realizados_dato.font = font_color_blanco_18
        celda_cantidad_realizados_dato.fill = fill_color_amarillo
        celda_cantidad_realizados_dato.border = border
        celda_cantidad_realizados_dato.alignment = aliniacion_central

        celda_cantidad_pendiente.font = font_color_negro_18
        celda_cantidad_pendiente.fill = fill_color_azul
        celda_cantidad_pendiente.border = border
        celda_cantidad_pendiente.alignment = aliniacion_derecha

        celda_cantidad_pendiente_dato.font = font_color_rojo_24
        celda_cantidad_pendiente_dato.fill = fill_color_amarillo
        celda_cantidad_pendiente_dato.border = border
        celda_cantidad_pendiente_dato.alignment = aliniacion_central

        celda_retornar.font = font_color_negro_18_nobold
        celda_retornar.border = border
        celda_retornar.alignment = aliniacion_centro_abajo

        celda_retornar_dato.font = font_color_blanco_24
        celda_retornar_dato.fill = fill_color_verde
        celda_retornar_dato.border = border
        celda_retornar_dato.alignment = aliniacion_centro_abajo
        celda_retornar_dato.number_format = moneda
        
        celda_gerome.font = font_color_negro_18
        celda_gerome.fill = fill_color_rosa
        celda_gerome.border = border
        celda_gerome.alignment = aliniacion_centro_abajo

        celda_numero.font = font_color_blanco_18_nobold
        celda_numero.fill = fill_color_azul
        celda_numero.border = border
        celda_numero.alignment = aliniacion_central

        celda_cuota_escrita.font = font_color_blanco_18_nobold
        celda_cuota_escrita.fill = fill_color_azul
        celda_cuota_escrita.border = border
        celda_cuota_escrita.alignment = aliniacion_central

        celda_fecha_escrita.font = font_color_blanco_18_nobold
        celda_fecha_escrita.fill = fill_color_azul
        celda_fecha_escrita.border = border
        celda_fecha_escrita.alignment = aliniacion_central

        celda_pagado_escrita.font = font_color_blanco_18_nobold
        celda_pagado_escrita.fill = fill_color_azul
        celda_pagado_escrita.border = border
        celda_pagado_escrita.alignment = aliniacion_central

        celda_balance_escrita.font = font_color_blanco_18_nobold
        celda_balance_escrita.fill = fill_color_azul
        celda_balance_escrita.border = border
        celda_balance_escrita.alignment = aliniacion_central
        
        celda_comentario_escrita.font = font_color_blanco_18_nobold
        celda_comentario_escrita.fill = fill_color_azul
        celda_comentario_escrita.border = border
        celda_comentario_escrita.alignment = aliniacion_central
        #Fin
        
        #Dar valor a las celdas
        celda_nombre.value = nombre

        celda_fecha.value = "Fecha inicio pr"
        celda_fecha_dato.value = fecha_inicio
        celda_gana.value = "Gana"
        celda_gana_dato.value = ganancia

        celda_monto.value = "Monto"
        celda_monto_dato.value = monto
        celda_cuota.value = "Cuota"
        celda_cuota_dato.value = cuota

        celda_cantidad_pagos.value = "Cantidad de pagos"
        celda_cantidad_pagos_dato.value = cantidad_pagos
        celda_modalidad.value = "Modalidad"
        celda_modalidad_dato.value = modal

        celda_cantidad_realizados.value = "Cant. Pag. Realizados"
        celda_cantidad_realizados_dato.value = "0"

        celda_retornar.value = "A RETONAR"
        celda_retornar_dato.value = retornar

        celda_cantidad_pendiente.value = "Cant. Pag. Pend."
        celda_cantidad_pendiente_dato.value = cantidad_pagos

        celda_gerome.value = "GEROME"

        celda_numero.value = "#"
        celda_cuota_escrita.value = "CUOTA"
        celda_fecha_escrita.value = "FECHA"
        celda_pagado_escrita.value = "PAGADO"
        celda_balance_escrita.value = "BALANCE"
        celda_comentario_escrita.value = "COMENTARIO"
        #Fin

        #Celdas de datos
        for x in range(6):
            for y in range(cantidad_pagos):
                palabra_celda = letras[x] + str(y + 10)
                hoja[palabra_celda].border = border
                hoja[palabra_celda].alignment = aliniacion_central

                if (letras[x] == 'B'):
                    hoja[palabra_celda].font = font_color_negro_12_nobold
                    hoja[palabra_celda].alignment = aliniacion_central
                    hoja[palabra_celda].value = y + 1
                elif (letras[x] == 'C'):
                    hoja[palabra_celda].font = font_color_negro_12_nobold
                    hoja[palabra_celda].number_format = moneda
                    hoja[palabra_celda].alignment = aliniacion_central
                    hoja[palabra_celda].value = cuota
                elif (letras[x] == 'D'):
                    hoja[palabra_celda].font = font_color_negro_12
                    hoja[palabra_celda].alignment = aliniacion_central
                elif (letras[x] == 'E'):
                    hoja[palabra_celda].font = font_color_negro_12
                    hoja[palabra_celda].alignment = aliniacion_central
                    hoja[palabra_celda].number_format = moneda
                elif (letras[x] == 'F'):
                    hoja[palabra_celda].font = font_color_negro_12
                    hoja[palabra_celda].alignment = aliniacion_central
                    hoja[palabra_celda].number_format = moneda
                elif (letras[x] == 'G'):
                    hoja[palabra_celda].font = font_color_negro_12
                    hoja[palabra_celda].alignment = aliniacion_central
        #Fin

        arch.save('C:/Users/Jimy/Desktop/Jatnna/' + nombre + '.xlsx')
    
    def agregar_datos(self, nombre, fecha, pagado):

        excel = openpyxl.load_workbook('C:/Users/Jimy/Desktop/Jatnna/' + nombre + '.xlsx')
        hoja_leer = excel.active

        cantidad_de_pagos = hoja_leer['C6']
        cantidad_de_pago_realizado = hoja_leer['C7']
        celda_cant_pagos_pendiente = hoja_leer['C8']
        celda_cuota = hoja_leer['E5']

        cantidad_de_pago_realizado.value = int(cantidad_de_pago_realizado.value) + 1

        indice = 9
        verificar = hoja_leer['E10'].value

        celda_cant_pagos_pendiente.value = cantidad_de_pagos.value - cantidad_de_pago_realizado.value

        for x in range(cantidad_de_pago_realizado.value):
            indice = indice + 1

        hoja_leer['D' + str (indice)].value = fecha
        hoja_leer['E' + str (indice)].value = pagado

        if (verificar is None):
            hoja_leer['F' + str (indice)].value = float(hoja_leer['E7'].value) - pagado

        else:
            hoja_leer['F' + str (indice)].value = float(hoja_leer['F' + str (indice - 1)].value) - pagado

        if (pagado >= celda_cuota.value):
            hoja_leer['G' + str (indice)].value = "PAGADO"

        else:
            hoja_leer['G' + str (indice)].value = "INCOMPLETO"

        excel.save('C:/Users/Jimy/Desktop/Jatnna/' + nombre + '.xlsx')

'''agregar_datos = crear_estado_de_cuenta()
agregar_datos.agregar_datos(nombre = "jimy", fecha = "15 AGOSTO", pagado = 8700.00)

crear = crear_estado_de_cuenta()
crear.crear(nombre="jimy", fecha_inicio = "30 DE JULIO 2021", cantidad_pagos = 16, modal = "SEMANAL", monto = 100000.00, cuota = 8700.00)'''